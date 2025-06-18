from django.db.models import Count, F, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt

from FDCSM import models
from FDCSM.utils.pagination import Pagination
from FDCSM.view.myModelForm import ProfessorModelForm, ProfessorSelectModelForm


def professor_info(request):
    # 管理员信息
    form = ProfessorModelForm()
    querySet = models.FDC_PFS_INFO.objects.all()
    page_object = Pagination(request, querySet)
    context = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, "professor_info.html", context)


#
@csrf_exempt
def professor_add(request):
    form = ProfessorModelForm(request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


@csrf_exempt
def professor_excel_add(request):
    from openpyxl import load_workbook

    file_object = request.FILES.get("PFSExcelFile")
    wb = load_workbook(file_object)
    ws = wb.active
    if ws is None:
        return JsonResponse({"status": False, "error": "Excel文件无有效工作表！"})
    if ws["A1"].value != "导师登录号码":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["B1"].value != "导师姓名":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["C1"].value != "导师登录密码":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["D1"].value != "导师简介":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["E1"].value != "物理专业可选人数":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["F1"].value != "数学专业可选人数":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["G1"].value != "系统科学专业可选人数":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["H1"].value != "应用统计专业可选人数":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["I1"].value != "纳米科学与工程专业可选人数":
        return JsonResponse({"status": False, "error": "模板格式错误！"})

    for row in ws.iter_rows(min_row=2):
        exists = models.FDC_PFS_INFO.objects.filter(PFS_NBR=row[0].value)
        if not exists:
            models.FDC_PFS_INFO.objects.create(
                PFS_NBR=row[0].value,
                PFS_NAM=row[1].value,
                PFS_PWD=row[2].value,
                PFS_INTRO=row[3].value,
                PFS_NUM_MATH=row[4].value if row[4].value else 0,
                PFS_NUM_PSC=row[5].value if row[5].value else 0,
                PFS_NUM_SCI=row[6].value if row[6].value else 0,
                PFS_NUM_AST=row[7].value if row[7].value else 0,  # 新增
                PFS_NUM_NANO=row[8].value if row[8].value else 0,  # 新增
            )
        else:
            exists.update(
                PFS_NAM=row[1].value,
                PFS_PWD=row[2].value,
                PFS_INTRO=row[3].value,
                PFS_NUM_MATH=row[4].value if row[4].value else 0,
                PFS_NUM_PSC=row[5].value if row[5].value else 0,
                PFS_NUM_SCI=row[6].value if row[6].value else 0,
                PFS_NUM_AST=row[7].value if row[7].value else 0,  # 新增
                PFS_NUM_NANO=row[8].value if row[8].value else 0,  # 新增
            )

    return JsonResponse({"status": True})


@csrf_exempt
def professor_excel_temp(request):
    import os

    from django.http import FileResponse

    file_path = os.path.join("FDCSM", "static", "excel", "20240421.xlsx")  # 文件路径
    file = open(file_path, "rb")  # 打开文件
    response = FileResponse(file)  # 创建FileResponse对象
    return response


def professor_detail(request):
    uid = request.GET.get("uid")
    row_dict = (
        models.FDC_PFS_INFO.objects.filter(PFS_NBR=uid)
        .values(
            "PFS_NBR",
            "PFS_NAM",
            "PFS_PWD",
            "PFS_TEL_NBR",
            "PFS_INTRO",
            "PFS_NUM_MATH",
            "PFS_NUM_PSC",
            "PFS_NUM_SCI",
            "PFS_NUM_AST",  # 新增
            "PFS_NUM_NANO",  # 新增
        )
        .first()
    )
    if not row_dict:
        return JsonResponse({"status": False, "error": "数据不存在。"})
    return JsonResponse({"status": True, "data": row_dict})


def professor_edit(request):
    uid = request.GET.get("uid")
    row_obj = models.FDC_PFS_INFO.objects.filter(PFS_NBR=uid).first()
    if not row_obj:
        return JsonResponse({"status": False, "tips": "数据不存在！"})
    form = ProfessorModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def professor_delete(request):
    uid = request.GET.get("uid")
    models.FDC_PFS_INFO.objects.filter(PFS_NBR=uid).delete()
    return JsonResponse({"status": True})


def professor_chose(request):
    uid = request.session.get("info")["id"]
    status = models.FDC_STAT_INFO.objects.all().filter().first()
    queryset1 = models.FDC_PFS_SEL.objects.filter(PFS_NBR=uid)
    form = ProfessorModelForm()
    if (status is not None) and (status.STAT_ID == 1):
        querySet = models.FDC_STU_INTRO.objects.filter(
            STU_ONE_SEL=uid, STU_NBR__STU_TYP=0
        )
    elif (status is not None) and (status.STAT_ID == 2):
        querySet = models.FDC_STU_INTRO.objects.filter(
            STU_TOW_SEL=uid, STU_NBR__STU_TYP=0
        )
    elif (status is not None) and (status.STAT_ID == 3):
        querySet = models.FDC_STU_INTRO.objects.filter(
            STU_THR_SEL=uid, STU_NBR__STU_TYP=0
        )
    else:
        return render(request, "professor_chose.html", {"form": form})
    page_object = Pagination(request, querySet)

    context = {
        "form": form,
        "queryset1": queryset1,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, "professor_chose.html", context)


@csrf_exempt
def professor_select(request):
    stu_id = request.POST.get("uid")
    stu_pro = request.POST.get("pro")
    pfs_id = request.session.get("info")["id"]
    # querySet = models.FDC_PFS_INFO.objects.annotate(select_count=Count('fdc_pfs_sel')).exclude(
    #     Q(PFS_NUM_PSC=0) | Q(select_count__gte=F('PFS_NUM_PSC')))
    pro_choice = {
        "0": "PFS_NUM_PSC",
        "1": "PFS_NUM_MATH",
        "2": "PFS_NUM_SCI",
        "3": "PFS_NUM_AST",
        "4": "PFS_NUM_NANO",
    }  # 添加新的专业
    status = models.FDC_STAT_INFO.objects.values("STAT_ID").all().first()
    if status is None or "STAT_ID" not in status:
        return JsonResponse(
            {"status": False, "errors": "系统状态信息不存在，无法选择学生"}
        )
    count_select = (
        models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=int(stu_pro)))
        )
        .filter(Q(select_count__lt=F(pro_choice[stu_pro])) & Q(PFS_NBR=pfs_id))
        .count()
    )
    if not count_select:
        return JsonResponse({"status": False, "errors": "该专业可选名额已满"})
    form = ProfessorSelectModelForm(
        {
            "PFS_NBR": pfs_id,
            "PFS_STU_NBR": stu_id,
            "PFS_STATE": status["STAT_ID"],
            "STU_PRO": stu_pro,
        }
    )
    if form.is_valid():
        form.save()
        models.FDC_STU_INFO.objects.filter(STU_NBR=stu_id).update(STU_TYP=1)
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "errors": form.errors})


@csrf_exempt
def professor_select_delect(request):
    stu_id = request.POST.get("uid")
    pfs_id = request.session.get("info")["id"]
    status = models.FDC_STAT_INFO.objects.values("STAT_ID").all().first()
    select_obj = models.FDC_PFS_SEL.objects.filter(
        PFS_NBR=pfs_id, PFS_STU_NBR=stu_id
    ).first()
    if (status is None) or (select_obj is None):
        return JsonResponse({"status": False, "errors": "系统状态或选择对象不存在"})
    if status["STAT_ID"] == select_obj.PFS_STATE:
        select_obj.delete()
        models.FDC_STU_INFO.objects.filter(STU_NBR=stu_id).update(STU_TYP=0)
        return JsonResponse({"status": True})
    else:
        return JsonResponse({"status": False, "errors": "不可删除其他批次的学生"})
