# FDCSM/view/myStudent.py

import django.utils.timezone as timezone
from django.db.models import Count, F, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt

from FDCSM import models
from FDCSM.utils.pagination import Pagination
from FDCSM.view.myModelForm import StudentModelForm, UploadModelForm


def student_info(request):
    # 管理员信息
    form = StudentModelForm()
    querySet = models.FDC_STU_INFO.objects.all()
    page_object = Pagination(request, querySet)
    context = {
        "form": form,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, "student_info.html", context)


@csrf_exempt
def student_add(request):
    form = StudentModelForm(request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


@csrf_exempt
def student_excel_add(request):
    from openpyxl import load_workbook

    file_object = request.FILES.get("STUExcelFile")
    wb = load_workbook(file_object)
    ws = wb.active
    if (ws is None) or (ws["A1"].value != "学生登录账号"):
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["B1"].value != "学生姓名":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if ws["C1"].value != "学生登录密码":
        return JsonResponse({"status": False, "error": "模板格式错误！"})
    if (
        ws["D1"].value
        != "学生专业（物理专业/数学专业/系统科学专业/应用统计专业/纳米科学与工程专业）"
    ):  # 更新表头
        return JsonResponse({"status": False, "error": "模板格式错误！"})

    stu_pro_choices = {
        "物理专业": 0,
        "数学专业": 1,
        "系统科学专业": 2,
        "应用统计专业": 3,
        "纳米科学与工程专业": 4,
    }  # 更新专业代码映射
    for row in ws.iter_rows(min_row=2):
        row_obj = models.FDC_STU_INFO.objects.filter(STU_NBR=row[0].value).first()
        data = {
            "STU_NBR": row[0].value,
            "STU_NAM": row[1].value,
            "STU_PWD": row[2].value,
            "STU_PRO": stu_pro_choices[str(row[3].value)],
        }
        if not row_obj:
            form = StudentModelForm(data)
            if form.is_valid():
                form.save()
            else:
                return JsonResponse({"status": False, "error": form.errors})
        else:
            form = StudentModelForm(data=data, instance=row_obj)
            if form.is_valid():
                form.save()
            else:
                return JsonResponse({"status": False, "error": form.errors})
    return JsonResponse({"status": True})


@csrf_exempt
def student_excel_temp(request):
    import os

    from django.http import FileResponse

    file_path = os.path.join("FDCSM", "static", "excel", "20240420.xlsx")  # 文件路径
    file = open(file_path, "rb")  # 打开文件
    response = FileResponse(file)  # 创建FileResponse对象
    return response


def student_delete(request):
    uid = request.GET.get("uid")
    models.FDC_STU_INFO.objects.filter(STU_NBR=uid).delete()
    return JsonResponse({"status": True})


def student_detail(request):
    uid = request.GET.get("uid")
    row_dict = (
        models.FDC_STU_INFO.objects.filter(STU_NBR=uid)
        .values("STU_NBR", "STU_NAM", "STU_PWD", "STU_TEL_NBR", "STU_PRO")
        .first()
    )
    if not row_dict:
        return JsonResponse({"status": False, "error": "数据不存在。"})
    return JsonResponse({"status": True, "data": row_dict})


def student_edit(request):
    uid = request.GET.get("uid")
    row_obj = models.FDC_STU_INFO.objects.filter(STU_NBR=uid).first()
    if not row_obj:
        return JsonResponse({"status": False, "tips": "数据不存在！"})
    form = StudentModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def student_chose(request):
    info_id = request.session.get("info")["id"]
    form = StudentModelForm()
    stu_info = models.FDC_STU_INFO.objects.get(STU_NBR=info_id)
    queryVit = models.FDC_STU_INTRO.objects.filter(STU_NBR=info_id).first()
    querySet = models.FDC_PFS_INFO.objects.none()  # 默认空QuerySet，防止未绑定

    if stu_info.STU_PRO == 0:  # 物理专业
        querySet = models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=0))
        ).exclude(Q(PFS_NUM_PSC=0) | Q(select_count__gte=F("PFS_NUM_PSC")))
    elif stu_info.STU_PRO == 1:  # 数学专业
        querySet = models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=1))
        ).exclude(Q(PFS_NUM_MATH=0) | Q(select_count__gte=F("PFS_NUM_MATH")))
    elif stu_info.STU_PRO == 2:  # 系统科学专业
        querySet = models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=2))
        ).exclude(Q(PFS_NUM_SCI=0) | Q(select_count__gte=F("PFS_NUM_SCI")))
    elif stu_info.STU_PRO == 3:  # 应用统计专业  # 新增
        querySet = models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=3))
        ).exclude(Q(PFS_NUM_AST=0) | Q(select_count__gte=F("PFS_NUM_AST")))
    elif stu_info.STU_PRO == 4:  # 纳米科学与工程专业  # 新增
        querySet = models.FDC_PFS_INFO.objects.annotate(
            select_count=Count("fdc_pfs_sel", Q(fdc_pfs_sel__STU_PRO=4))
        ).exclude(Q(PFS_NUM_NANO=0) | Q(select_count__gte=F("PFS_NUM_NANO")))
    page_object = Pagination(request, querySet)
    context = {
        "form": form,
        "queryVit": queryVit,
        # 'stu_sel_obj': stu_sel_obj,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, "student_chose.html", context)


@csrf_exempt
def student_select(request):
    teacher_id = request.POST.get("uid")
    info_id = request.session.get("info")["id"]
    stu_sel_obj = models.FDC_STU_INTRO.objects.filter(STU_NBR=info_id)
    stu_info = models.FDC_STU_INFO.objects.filter(STU_NBR=info_id)
    stu_info_obj = stu_info.first()
    if stu_info_obj and hasattr(stu_info_obj, "STU_TYP") and stu_info_obj.STU_TYP == 1:
        return JsonResponse(
            {"status": False, "error": "你已经被导师录用，不用再选择导师"}
        )
    if stu_sel_obj:
        if not getattr(stu_sel_obj.first(), "STU_ONE_SEL_id", None):
            stu_sel_obj.update(STU_ONE_SEL_id=teacher_id, CHG_TIM=timezone.now())
        elif not getattr(stu_sel_obj.first(), "STU_TOW_SEL_id", None):
            stu_sel_obj.update(STU_TOW_SEL_id=teacher_id, CHG_TIM=timezone.now())
        elif not getattr(stu_sel_obj.first(), "STU_THR_SEL_id", None):
            stu_sel_obj.update(STU_THR_SEL_id=teacher_id, CHG_TIM=timezone.now())
        else:
            return JsonResponse({"status": False, "error": "不能再选导师"})
    else:
        models.FDC_STU_INTRO.objects.create(
            STU_NBR_id=info_id, STU_ONE_SEL_id=teacher_id
        )
    return JsonResponse({"status": True})


def student_delete_sel(request):
    info_id = request.session.get("info")["id"]
    obj = models.FDC_STU_INTRO.objects.filter(STU_NBR=info_id)
    obj.update(
        STU_ONE_SEL=None, STU_TOW_SEL=None, STU_THR_SEL=None, CHG_TIM=timezone.now()
    )
    return redirect("/student_info/student_chose/")


def student_upload(request):
    """上传简历和成绩"""
    stu_id = request.session.get("info")["id"]
    if request.method == "GET":
        form = UploadModelForm()
        return render(request, "upload_form.html", {"form": form})
    instance = models.FDC_STU_INTRO.objects.filter(STU_NBR=stu_id).first()
    print(instance)
    if instance:
        form = UploadModelForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
    else:
        form = UploadModelForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.STU_NBR_id = stu_id
            form.save()
    return redirect("/student_info/upload/")
    # return render(request, 'upload_form.html', {'form': form})
